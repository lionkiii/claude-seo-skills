#!/usr/bin/env python3
"""
Build an interactive multi-sheet SEO audit workbook from crawl_results.json
plus optional GSC and Ahrefs data files.

Every actionable row gets a Status dropdown (To Do / In Progress / Done /
Skipped / Not Applicable) with conditional row colouring:
    Done             -> green
    Skipped / N/A    -> grey
    In Progress      -> soft yellow

Usage:
    python build_audit_xlsx.py \\
        --crawl /tmp/audit/crawl_results.json \\
        --out ~/Desktop/seo-audit-example.com-2026-04-24.xlsx \\
        --site-name "example.com" \\
        [--gsc-top-pages /tmp/audit/gsc_top_pages.json] \\
        [--gsc-queries /tmp/audit/gsc_queries.json] \\
        [--gsc-devices /tmp/audit/gsc_devices.json] \\
        [--ahrefs-metrics /tmp/audit/ahrefs_metrics.json] \\
        [--ahrefs-top-pages /tmp/audit/ahrefs_top_pages.json] \\
        [--ahrefs-keywords /tmp/audit/ahrefs_keywords.json] \\
        [--dr 92] \\
        [--brand-regex "zoho|campaing|..."]

Data-file formats: each is a JSON list-of-dicts or dict with a known key.
See skills/seo-audit/scripts/README.md for examples.
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl required. Install with: pip install 'openpyxl>=3.1,<4.0'", file=sys.stderr)
    sys.exit(1)


# ---------------------------- styles ----------------------------------------

TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", size=11, bold=True)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_CRIT = PatternFill("solid", fgColor="C00000")
FILL_HIGH = PatternFill("solid", fgColor="ED7D31")
FILL_MED = PatternFill("solid", fgColor="FFC000")
FILL_LOW = PatternFill("solid", fgColor="A9D08E")
FILL_DONE = PatternFill("solid", fgColor="C6EFCE")
FILL_SKIP = PatternFill("solid", fgColor="D9D9D9")
FILL_PROG = PatternFill("solid", fgColor="FFF2CC")
FILL_TODO = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center")

STATUS_VALUES = '"To Do,In Progress,Done,Skipped,Not Applicable"'
PRIORITY_VALUES = '"Critical,High,Medium,Low"'
EFFORT_VALUES = '"Easy,Medium,Hard"'

PRI_FILL = {"Critical": FILL_CRIT, "High": FILL_HIGH, "Medium": FILL_MED, "Low": FILL_LOW}
PRI_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}


def pri_font(pri: str) -> Font:
    dark = pri in ("Critical", "High")
    return Font(color="FFFFFF" if dark else "000000", bold=True)


# ---------------------------- sheet helpers ---------------------------------

def add_title_row(ws, text: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28


def add_header_row(ws, row: int, headers: list) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_list_validation(ws, col_letter: str, start: int, end: int, values_csv: str) -> None:
    dv = DataValidation(type="list", formula1=values_csv, allow_blank=True)
    dv.add(f"{col_letter}{start}:{col_letter}{end}")
    ws.add_data_validation(dv)


def apply_status_cf(ws, status_col: str, start: int, end: int, first_col: str, last_col: str) -> None:
    rng = f"{first_col}{start}:{last_col}{end}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'${status_col}{start}="Done"'], fill=FILL_DONE))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'OR(${status_col}{start}="Skipped",${status_col}{start}="Not Applicable")'],
        fill=FILL_SKIP))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'${status_col}{start}="In Progress"'], fill=FILL_PROG))


# ---------------------------- scoring ---------------------------------------

WEIGHTS = {"tech": 0.25, "content": 0.25, "onpage": 0.20, "schema": 0.10,
           "perf": 0.10, "img": 0.05, "ai": 0.05}


def compute_scores(ok_pages: list[dict]) -> dict:
    """Rough per-category scoring driven by crawl signals."""
    n = max(1, len(ok_pages))
    no_schema = sum(1 for p in ok_pages if p["jsonld_blocks"] == 0)
    no_hreflang = sum(1 for p in ok_pages if p["hreflang_count"] == 0)
    h1_bad = sum(1 for p in ok_pages if p["h1_count"] != 1)
    title_bad = sum(1 for p in ok_pages if p["title_len"] == 0 or p["title_len"] < 30 or p["title_len"] > 65)
    meta_bad = sum(1 for p in ok_pages if p["meta_desc_len"] == 0 or p["meta_desc_len"] < 120 or p["meta_desc_len"] > 160)
    thin = sum(1 for p in ok_pages if p["word_count"] < 300)
    missing_alt_pages = sum(1 for p in ok_pages if p["images_missing_alt"] > 0)

    schema = int(100 * (1 - no_schema / n))
    tech_penalty = (no_hreflang / n) * 20
    tech = max(50, int(100 - tech_penalty))
    onpage = int(100 * (1 - (title_bad + meta_bad + h1_bad) / (3 * n)))
    content = int(100 * (1 - thin / n * 1.2))
    img = int(100 * (1 - missing_alt_pages / n))
    perf = 70  # default when no CWV sample
    ai = int((schema * 0.6) + (40 if thin / n < 0.3 else 20))

    overall = int(round(
        tech * WEIGHTS["tech"] + content * WEIGHTS["content"] + onpage * WEIGHTS["onpage"]
        + schema * WEIGHTS["schema"] + perf * WEIGHTS["perf"] + img * WEIGHTS["img"]
        + ai * WEIGHTS["ai"]))
    return {
        "overall": overall, "tech": tech, "content": max(0, content),
        "onpage": max(0, onpage), "schema": schema, "perf": perf,
        "img": max(0, img), "ai": max(0, ai),
    }


# ---------------------------- page-type classifier --------------------------

def infer_page_type(url: str) -> str:
    u = url.lower()
    host_path = u.split("://", 1)[-1]
    path = "/" + host_path.split("/", 1)[1] if "/" in host_path else "/"
    if path in ("/", ""): return "Homepage"
    if "pricing" in u: return "Pricing"
    if "/help/" in u or "/docs/" in u or "/developers/" in u or "/api/" in u or "/emailapi/" in u:
        return "Help / API docs"
    if "/webinar" in u or "/events/" in u: return "Webinar / Event"
    if "/customer" in u or "/case-stud" in u: return "Case study"
    if "/blog/" in u or "/guides/" in u or "/learn/" in u or "marketingmatchbox" in u:
        return "Blog / guide"
    if "/integration" in u or "/integrations" in u: return "Integration"
    if "-alternative" in u or "-vs-" in u or "-comparison" in u: return "Competitor comparison"
    if "/features" in u or "-feature" in u: return "Feature page"
    if u.endswith("login.html") or u.endswith("signup.html") or "/login" in u or "/signup" in u:
        return "Utility (login/signup)"
    if "/solutions/" in u or "/industr" in u: return "Solution / industry"
    if "/explainer/" in u: return "Explainer"
    return "Content / product page"


def recommend_schema(ptype: str) -> str:
    return {
        "Homepage": "Organization + WebSite (with SearchAction) + SoftwareApplication",
        "Pricing": "Product + Offer + FAQPage (if Q&As) + BreadcrumbList",
        "Help / API docs": "TechArticle + BreadcrumbList",
        "Webinar / Event": "Event + VideoObject + BreadcrumbList",
        "Case study": "Article + Organization (customer) + BreadcrumbList",
        "Blog / guide": "Article + BreadcrumbList + Person (author)",
        "Integration": "SoftwareApplication + BreadcrumbList",
        "Competitor comparison": "Article + FAQPage + BreadcrumbList",
        "Feature page": "SoftwareApplication feature + BreadcrumbList",
        "Utility (login/signup)": "Consider noindex OR WebPage + BreadcrumbList",
        "Solution / industry": "Service + BreadcrumbList",
        "Explainer": "Article + BreadcrumbList",
        "Content / product page": "WebPage + BreadcrumbList",
    }.get(ptype, "WebPage + BreadcrumbList")


# ---------------------------- sheet builders --------------------------------

REF_TITLE = "https://developers.google.com/search/docs/appearance/title-link"
REF_META = "https://developers.google.com/search/docs/appearance/snippet"
REF_H1 = "https://web.dev/articles/heading-order"
REF_SCHEMA = "https://developers.google.com/search/docs/appearance/structured-data/intro-structured-data"
REF_HREFLANG = "https://developers.google.com/search/docs/specialty/international/localized-versions"
REF_CONTENT = "https://developers.google.com/search/docs/fundamentals/creating-helpful-content"
REF_MOBILE = "https://developers.google.com/search/docs/crawling-indexing/mobile/mobile-sites-mobile-first-indexing"

REFS_TABLE = [
    ("SEO fundamentals (Starter Guide)", "https://developers.google.com/search/docs/fundamentals/seo-starter-guide"),
    ("Creating helpful content (E-E-A-T)", REF_CONTENT),
    ("Title link best practices", REF_TITLE),
    ("Meta description / snippet best practices", REF_META),
    ("Intro to structured data (JSON-LD)", REF_SCHEMA),
    ("Organization schema", "https://developers.google.com/search/docs/appearance/structured-data/organization"),
    ("Product schema", "https://developers.google.com/search/docs/appearance/structured-data/product"),
    ("SoftwareApplication schema", "https://developers.google.com/search/docs/appearance/structured-data/software-app"),
    ("FAQPage schema", "https://developers.google.com/search/docs/appearance/structured-data/faqpage"),
    ("BreadcrumbList schema", "https://developers.google.com/search/docs/appearance/structured-data/breadcrumb"),
    ("Article schema", "https://developers.google.com/search/docs/appearance/structured-data/article"),
    ("VideoObject schema", "https://developers.google.com/search/docs/appearance/structured-data/video"),
    ("Event schema", "https://developers.google.com/search/docs/appearance/structured-data/event"),
    ("Sitelinks search box (WebSite)", "https://developers.google.com/search/docs/appearance/structured-data/sitelinks-searchbox"),
    ("Internationalization / hreflang", REF_HREFLANG),
    ("Canonical tag guide", "https://developers.google.com/search/docs/crawling-indexing/canonicalization"),
    ("Sitemaps overview", "https://developers.google.com/search/docs/crawling-indexing/sitemaps/overview"),
    ("robots.txt introduction", "https://developers.google.com/search/docs/crawling-indexing/robots/intro"),
    ("Mobile-first indexing", REF_MOBILE),
    ("Core Web Vitals (web.dev)", "https://web.dev/articles/vitals"),
    ("PageSpeed Insights tool", "https://pagespeed.web.dev/"),
    ("Rich Results Test", "https://search.google.com/test/rich-results"),
    ("Schema Markup Validator", "https://validator.schema.org/"),
    ("Google Search Central docs", "https://developers.google.com/search/docs"),
    ("LLMs.txt specification", "https://llmstxt.org/"),
    ("Heading order (web.dev)", REF_H1),
    ("Google Images best practices", "https://developers.google.com/search/docs/appearance/google-images"),
]


def build_read_me(wb, site: str, n_pages: int) -> None:
    ws = wb.create_sheet("Read Me")
    add_title_row(ws, f"SEO Audit — {site}", 1)
    ws.column_dimensions["A"].width = 120
    lines = [
        "",
        "AUDIT SCOPE",
        f"  Target: {site}",
        f"  Pages analysed: {n_pages}",
        f"  Date: {datetime.now().strftime('%Y-%m-%d')}",
        "  Data sources: (1) Live crawl, (2) Google Search Console [if supplied], (3) Ahrefs [if supplied]",
        "",
        "HOW TO USE",
        "  1. Start with 'Executive Summary' — score + top issues + top wins.",
        "  2. 'Action Plan' is the master to-do list. Status dropdown: To Do / In Progress / Done / Skipped / N/A.",
        "  3. Rows turn GREEN when Done, GREY when Skipped/N/A, soft YELLOW when In Progress.",
        "  4. Use detail tabs (Schema Missing, Thin Content, On-Page Issues...) for deeper per-page context.",
        "  5. 'Google References' tab lists every cited authority document.",
        "",
        "PRIORITY LEVELS",
        "  CRITICAL  — Blocks indexing, rich results, or causes penalties. Fix immediately.",
        "  HIGH      — Significantly impacts rankings / CTR. Fix within 1 week.",
        "  MEDIUM    — Optimisation opportunity. Fix within 1 month.",
        "  LOW       — Nice to have. Backlog.",
        "",
        "EFFORT LEVELS",
        "  EASY    — Under 1 hour (single-file edit, copy change).",
        "  MEDIUM  — A few hours to a day (template change, schema generator).",
        "  HARD    — Multi-day / cross-team (rewrites, performance work).",
    ]
    for i, L in enumerate(lines, 2):
        c = ws.cell(row=i, column=1, value=L)
        c.alignment = Alignment(horizontal="left", vertical="top")
        if L and not L.startswith(" "):
            c.font = BOLD


def build_executive_summary(wb, site: str, ok: list[dict], scores: dict, ahrefs_metrics: dict | None,
                             gsc_totals: dict | None, device_split: list | None) -> None:
    ws = wb.create_sheet("Executive Summary")
    add_title_row(ws, f"Executive Summary — {site} ({datetime.now().strftime('%Y-%m-%d')})", 3)
    set_col_widths(ws, [50, 22, 90])

    n = max(1, len(ok))
    rows: list[tuple] = [
        ("Overall SEO Health Score", f"{scores['overall']}/100", "Weighted across 7 categories."),
        ("Pages audited", len(ok), ""),
    ]
    if ahrefs_metrics:
        rows += [
            ("Ahrefs Organic Keywords", ahrefs_metrics.get("org_keywords", "n/a"), ""),
            ("Ahrefs Organic Traffic (monthly)", ahrefs_metrics.get("org_traffic", "n/a"), ""),
        ]
    if gsc_totals:
        rows += [
            ("GSC 90-day Clicks", f"{gsc_totals.get('clicks', 0):,}", ""),
            ("GSC 90-day Impressions", f"{gsc_totals.get('impressions', 0):,}", ""),
            ("GSC 90-day Avg CTR", f"{gsc_totals.get('ctr', 0):.2f}%", ""),
            ("GSC 90-day Avg Position", f"{gsc_totals.get('position', 0):.1f}", ""),
        ]
    if device_split:
        d_str = ", ".join(f"{d['device']}: pos {d['position']:.1f} / CTR {d['ctr']:.2f}%" for d in device_split)
        rows.append(("GSC Device Split", d_str, "Large desktop-vs-mobile position gap = mobile SEO issue."))
    rows += [
        ("", "", ""),
        ("CATEGORY SCORES", "", ""),
        ("Technical SEO (25%)", f"{scores['tech']}/100", "HTTPS/headers/robots. Weak: missing schema & hreflang in HTML."),
        ("Content Quality (25%)", f"{scores['content']}/100", "Driven by thin pages & duplicates."),
        ("On-Page SEO (20%)", f"{scores['onpage']}/100", "Title / meta / H1 coverage."),
        ("Schema / Structured Data (10%)", f"{scores['schema']}/100", "Rich-result eligibility."),
        ("Performance (CWV) (10%)", f"{scores['perf']}/100", "Run PageSpeed Insights for precise numbers."),
        ("Images (5%)", f"{scores['img']}/100", "Alt-text coverage."),
        ("AI Search Readiness (5%)", f"{scores['ai']}/100", "Schema + llms.txt + content depth."),
    ]

    no_schema = sum(1 for p in ok if p["jsonld_blocks"] == 0)
    no_hreflang = sum(1 for p in ok if p["hreflang_count"] == 0)
    thin = sum(1 for p in ok if p["word_count"] < 300)
    h1_bad = sum(1 for p in ok if p["h1_count"] != 1)
    rows += [
        ("", "", ""),
        ("TOP CRITICAL FINDINGS", "", ""),
        (f"{no_schema}/{n} pages have no JSON-LD schema", "CRITICAL" if no_schema / n > 0.5 else "HIGH",
         "Add Organization + WebSite schema globally, then per-page types (Product, FAQPage, BreadcrumbList)."),
        (f"{no_hreflang}/{n} pages have no hreflang in HTML", "CRITICAL" if no_hreflang / n > 0.5 else "HIGH",
         "Mirror sitemap hreflang into HTML <head>."),
        (f"{thin}/{n} pages have under 300 words", "HIGH", "Expand thin utility pages or noindex them."),
        (f"{h1_bad}/{n} pages have missing or multiple H1s", "HIGH", "Exactly one <h1> per page."),
    ]

    for r, (a, b, c_) in enumerate(rows, 3):
        ra = ws.cell(row=r, column=1, value=a); ra.font = BOLD if a and a.isupper() else BODY
        rb = ws.cell(row=r, column=2, value=b); rb.font = BOLD
        rc = ws.cell(row=r, column=3, value=c_); rc.font = BODY; rc.alignment = WRAP
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).border = BORDER
        if a == "Overall SEO Health Score":
            rb.fill = FILL_MED; rb.font = Font(name="Calibri", size=16, bold=True)
        elif isinstance(b, str) and b == "CRITICAL":
            rb.fill = FILL_CRIT; rb.font = pri_font("Critical")
        elif isinstance(b, str) and b == "HIGH":
            rb.fill = FILL_HIGH; rb.font = pri_font("High")


def build_action_plan(wb, ok: list[dict], crawl_meta: dict) -> None:
    ws = wb.create_sheet("Action Plan")
    headers = ["#", "Status", "Priority", "Effort", "Category", "Issue", "Pages affected",
               "How to fix", "Owner", "Google Reference"]
    add_title_row(ws, "Action Plan — Master To-Do List", len(headers))
    add_header_row(ws, 2, headers)

    n = max(1, len(ok))
    no_schema = sum(1 for p in ok if p["jsonld_blocks"] == 0)
    no_hreflang = sum(1 for p in ok if p["hreflang_count"] == 0)
    no_h1 = sum(1 for p in ok if p["h1_count"] == 0)
    multi_h1 = sum(1 for p in ok if p["h1_count"] > 1)
    t_long = sum(1 for p in ok if p["title_len"] > 65)
    t_short = sum(1 for p in ok if 0 < p["title_len"] < 30)
    m_short = sum(1 for p in ok if 0 < p["meta_desc_len"] < 120)
    m_long = sum(1 for p in ok if p["meta_desc_len"] > 160)
    m_none = sum(1 for p in ok if p["meta_desc_len"] == 0)
    thin = sum(1 for p in ok if p["word_count"] < 300)
    alt_bad = sum(1 for p in ok if p["images_missing_alt"] > 0)
    no_viewport = sum(1 for p in ok if not p["has_viewport"])

    actions: list[tuple] = []
    if no_schema:
        actions.append(("Critical", "Medium", "Schema",
                        f"No JSON-LD schema on {no_schema}/{n} pages",
                        f"{no_schema} pages",
                        "Add Organization + WebSite schema globally. Per-page: Product/FAQPage/BreadcrumbList/Article as fits page type. See 'Schema Missing' tab for per-URL recommendation.",
                        "Dev", REF_SCHEMA))
    if no_hreflang:
        actions.append(("Critical" if no_hreflang / n > 0.5 else "High", "Medium", "Hreflang",
                        f"No hreflang tags in HTML on {no_hreflang}/{n} pages",
                        f"{no_hreflang} pages",
                        "Emit <link rel='alternate' hreflang='x'> tags in <head> for every locale variant. Google recommends dual declaration (sitemap + HTML).",
                        "Dev", REF_HREFLANG))
    if no_h1:
        actions.append(("High", "Easy", "On-page",
                        f"{no_h1} pages missing H1 tag", f"{no_h1} pages",
                        "Every page must have exactly one <h1>. Add descriptive H1 matching primary keyword.",
                        "Marketing+Dev", REF_H1))
    if multi_h1:
        actions.append(("High", "Medium", "On-page",
                        f"{multi_h1} pages with multiple H1 tags", f"{multi_h1} pages",
                        "Use <h1> exactly once. Downgrade extras to <h2> or <h3>.",
                        "Dev", REF_H1))
    if t_long:
        actions.append(("High", "Easy", "Titles",
                        f"{t_long} titles exceed 65 characters (truncate in SERPs)", f"{t_long} pages",
                        "Shorten to ≤60 chars. Format: 'Primary Keyword — Benefit | Brand'.",
                        "Marketing", REF_TITLE))
    if t_short:
        actions.append(("High", "Easy", "Titles",
                        f"{t_short} titles under 30 characters", f"{t_short} pages",
                        "Expand to 50-60 chars with primary keyword + value prop.",
                        "Marketing", REF_TITLE))
    if m_none:
        actions.append(("High", "Easy", "Meta desc",
                        f"{m_none} pages missing meta description", f"{m_none} pages",
                        "Write 120-155 char description with keyword + CTA.",
                        "Marketing", REF_META))
    if m_short:
        actions.append(("High", "Easy", "Meta desc",
                        f"{m_short} pages with meta description <120 chars", f"{m_short} pages",
                        "Expand to 120-155 chars.", "Marketing", REF_META))
    if m_long:
        actions.append(("High", "Easy", "Meta desc",
                        f"{m_long} pages with meta description >160 chars", f"{m_long} pages",
                        "Trim to ≤155 chars.", "Marketing", REF_META))
    if thin:
        actions.append(("High", "Medium", "Content",
                        f"{thin} pages under 300 words (thin content)", f"{thin} pages",
                        "Expand with FAQs, related links, screenshots. Noindex bare shell pages.",
                        "Marketing", REF_CONTENT))
    if alt_bad:
        actions.append(("Medium", "Easy", "Images",
                        f"{alt_bad} pages with images missing alt", f"{alt_bad} pages",
                        "Add descriptive alt text. Use alt='' for purely decorative.",
                        "Marketing", "https://developers.google.com/search/docs/appearance/google-images"))
    if no_viewport:
        actions.append(("High", "Easy", "Mobile",
                        f"{no_viewport} pages missing viewport meta", f"{no_viewport} pages",
                        "Add <meta name='viewport' content='width=device-width, initial-scale=1'>.",
                        "Dev", REF_MOBILE))

    actions.extend([
        ("Medium", "Medium", "Performance",
         "Core Web Vitals not sampled in this audit", "Top 10 pages",
         "Run PageSpeed Insights on top 10 pages. Targets: LCP ≤2.5s, INP ≤200ms, CLS ≤0.1 at 75th pct mobile.",
         "Dev", "https://web.dev/articles/vitals"),
        ("Medium", "Medium", "AI Search",
         "Create llms.txt file", "Root domain",
         "Create /llms.txt listing key pages for LLM crawlers. Improves ChatGPT / Perplexity citation accuracy.",
         "Dev", "https://llmstxt.org/"),
    ])

    actions.sort(key=lambda a: PRI_ORDER[a[0]])
    start = 3
    for i, (pri, eff, cat, issue, pages, how, owner, ref) in enumerate(actions):
        r = start + i
        vals = [i + 1, "To Do", pri, eff, cat, issue, pages, how, owner, ref]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = WRAP; c.border = BORDER; c.font = BODY
        pc = ws.cell(row=r, column=3); pc.fill = PRI_FILL[pri]; pc.font = pri_font(pri)
        ws.cell(row=r, column=2).fill = FILL_TODO
        ws.row_dimensions[r].height = 70

    end = start + len(actions) - 1
    if actions:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_list_validation(ws, "C", start, end, PRIORITY_VALUES)
        apply_list_validation(ws, "D", start, end, EFFORT_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 15, 11, 9, 14, 48, 22, 65, 14, 55])
    ws.freeze_panes = "C3"


def build_page_inventory(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("Page Inventory")
    headers = ["#", "URL", "Status", "Title", "Title len", "Meta desc len", "H1 count", "H2 count",
               "Word count", "Hreflang", "Schema blocks", "Images", "Missing alt", "Canonical OK?"]
    add_title_row(ws, f"Page Inventory — {len(ok)} URLs", len(headers))
    add_header_row(ws, 2, headers)

    start = 3
    for i, p in enumerate(ok):
        r = start + i
        canon_ok = ("Yes" if p["canonical_self"] or p["canonical"].rstrip("/") == p["url"].rstrip("/")
                    else ("No" if p["canonical"] else "Missing"))
        vals = [i + 1, p["url"], "To Do", p["title"][:120], p["title_len"], p["meta_desc_len"],
                p["h1_count"], p["h2_count"], p["word_count"], p["hreflang_count"],
                p["jsonld_blocks"], p["images"], p["images_missing_alt"], canon_ok]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER; c.font = BODY
            c.alignment = WRAP if col in (2, 4) else CENTER
        # flags
        if p["title_len"] == 0 or p["title_len"] < 30 or p["title_len"] > 65:
            ws.cell(row=r, column=5).fill = FILL_MED
        if p["meta_desc_len"] == 0 or p["meta_desc_len"] < 120 or p["meta_desc_len"] > 160:
            ws.cell(row=r, column=6).fill = FILL_MED
        if p["h1_count"] != 1:
            ws.cell(row=r, column=7).fill = FILL_HIGH; ws.cell(row=r, column=7).font = pri_font("High")
        if p["word_count"] < 300:
            ws.cell(row=r, column=9).fill = FILL_MED
        if p["hreflang_count"] == 0:
            ws.cell(row=r, column=10).fill = FILL_CRIT; ws.cell(row=r, column=10).font = pri_font("Critical")
        if p["jsonld_blocks"] == 0:
            ws.cell(row=r, column=11).fill = FILL_CRIT; ws.cell(row=r, column=11).font = pri_font("Critical")
        if p["images_missing_alt"] > 0:
            ws.cell(row=r, column=13).fill = FILL_HIGH; ws.cell(row=r, column=13).font = pri_font("High")
    end = start + len(ok) - 1
    apply_list_validation(ws, "C", start, end, STATUS_VALUES)
    apply_status_cf(ws, "C", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 70, 14, 60, 10, 11, 9, 9, 11, 11, 13, 8, 11, 14])
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{end}"


def build_on_page_issues(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("On-Page Issues")
    headers = ["#", "Status", "Priority", "URL", "Issue type", "Current value", "Detail", "Recommended fix", "Reference"]
    add_title_row(ws, "On-Page Issues — Title, Meta, H1 per page", len(headers))
    add_header_row(ws, 2, headers)

    rows = []
    for p in ok:
        u = p["url"]
        if p["title_len"] == 0:
            rows.append(("Critical", u, "Title missing", "(empty)", "", "Add unique 50-60 char title.", REF_TITLE))
        elif p["title_len"] < 30:
            rows.append(("High", u, "Title too short", p["title"], f"{p['title_len']} chars", "Expand to 50-60 chars with primary keyword.", REF_TITLE))
        elif p["title_len"] > 65:
            rows.append(("High", u, "Title too long", p["title"], f"{p['title_len']} chars — truncated in SERPs", "Shorten to ≤60 chars.", REF_TITLE))
        if "&#039;" in p["title"] or "&amp;" in p["title"]:
            rows.append(("Low", u, "Title has HTML entity", p["title"], "encoded char", "Replace entity with real character in CMS.", REF_TITLE))
        if p["meta_desc_len"] == 0:
            rows.append(("Critical", u, "Meta desc missing", "(empty)", "", "Write 120-155 char description.", REF_META))
        elif p["meta_desc_len"] < 120:
            rows.append(("Medium", u, "Meta desc too short", p["meta_desc"], f"{p['meta_desc_len']} chars", "Expand to 120-155 chars.", REF_META))
        elif p["meta_desc_len"] > 160:
            rows.append(("Medium", u, "Meta desc too long", p["meta_desc"][:120] + "…", f"{p['meta_desc_len']} chars — truncated", "Trim to ≤155 chars.", REF_META))
        if p["h1_count"] == 0:
            rows.append(("High", u, "Missing H1", "(none)", "", "Add a single descriptive <h1>.", REF_H1))
        elif p["h1_count"] > 1:
            rows.append(("Medium", u, "Multiple H1s", p["h1_text"][:80], f"{p['h1_count']} H1s", "Keep only ONE <h1>.", REF_H1))

    rows.sort(key=lambda x: (PRI_ORDER[x[0]], x[1]))
    start = 3
    for i, (pri, url, itype, cur, det, fix, ref) in enumerate(rows):
        r = start + i
        vals = [i + 1, "To Do", pri, url, itype, cur, det, fix, ref]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.alignment = WRAP; c.font = BODY
        pc = ws.cell(row=r, column=3); pc.fill = PRI_FILL[pri]; pc.font = pri_font(pri)
    end = start + len(rows) - 1
    if rows:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_list_validation(ws, "C", start, end, PRIORITY_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 14, 11, 60, 22, 55, 25, 65, 55])
    ws.freeze_panes = "D3"
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{end}"


def build_schema_missing(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("Schema Missing")
    headers = ["#", "Status", "URL", "Page type (inferred)", "Recommended schema type", "Priority", "Google Reference"]
    missing = [p for p in ok if p["jsonld_blocks"] == 0]
    add_title_row(ws, f"Schema Missing — {len(missing)} pages with 0 JSON-LD", len(headers))
    add_header_row(ws, 2, headers)

    start = 3
    for i, p in enumerate(missing):
        r = start + i
        t = infer_page_type(p["url"])
        sch = recommend_schema(t)
        pri = ("Critical" if t in ("Homepage", "Pricing", "Feature page")
               else "High" if t in ("Help / API docs", "Webinar / Event", "Blog / guide", "Integration",
                                    "Competitor comparison", "Case study") else "Medium")
        vals = [i + 1, "To Do", p["url"], t, sch, pri, REF_SCHEMA]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
        pc = ws.cell(row=r, column=6); pc.fill = PRI_FILL[pri]; pc.font = pri_font(pri)
    end = start + len(missing) - 1
    if missing:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 14, 70, 22, 50, 10, 55])
    ws.freeze_panes = "C3"
    if missing:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{end}"


def build_hreflang_missing(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("Hreflang Missing")
    headers = ["#", "Status", "URL", "Hreflang in sitemap?", "Google Reference"]
    missing = [p for p in ok if p["hreflang_count"] == 0]
    add_title_row(ws, f"Hreflang Missing from HTML — {len(missing)} pages", len(headers))
    add_header_row(ws, 2, headers)

    note = ("NOTE: Hreflang may be in sitemap but not emitted in HTML <head>. "
            "Google accepts both; dual declaration is strongly recommended for reliability.")
    ws.cell(row=2, column=5).comment = Comment(note, "audit")

    start = 3
    for i, p in enumerate(missing):
        r = start + i
        vals = [i + 1, "To Do", p["url"], "Check sitemap", REF_HREFLANG]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
    end = start + len(missing) - 1
    if missing:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 14, 75, 25, 60])
    ws.freeze_panes = "C3"


def build_thin_content(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("Thin Content")
    headers = ["#", "Status", "Priority", "URL", "Word count", "Suggested action", "Google Reference"]
    thin = sorted([p for p in ok if p["word_count"] < 300], key=lambda x: x["word_count"])
    add_title_row(ws, f"Thin Content — {len(thin)} pages under 300 words", len(headers))
    add_header_row(ws, 2, headers)

    start = 3
    for i, p in enumerate(thin):
        r = start + i
        wc = p["word_count"]
        if wc < 50:
            pri, action = "Critical", "Bare page — noindex OR rebuild with substantial content + internal links."
        elif wc < 100:
            pri, action = "High", "Expand to 400+ words with FAQs, related links, product context."
        elif wc < 200:
            pri, action = "High", "Expand to 400+ words. Add use cases, captions, related content block."
        else:
            pri, action = "Medium", "Expand to 400+ words. Audit for duplicate / boilerplate content."
        vals = [i + 1, "To Do", pri, p["url"], wc, action, REF_CONTENT]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
        pc = ws.cell(row=r, column=3); pc.fill = PRI_FILL[pri]; pc.font = pri_font(pri)
    end = start + len(thin) - 1
    if thin:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_list_validation(ws, "C", start, end, PRIORITY_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 14, 11, 70, 12, 65, 55])
    ws.freeze_panes = "D3"


def build_duplicate_titles_meta(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("Duplicate Titles & Meta")
    headers = ["#", "Status", "Priority", "Type", "Shared content", "Affected URLs", "Fix"]
    add_title_row(ws, "Duplicate Titles & Meta Descriptions", len(headers))
    add_header_row(ws, 2, headers)

    title_groups, meta_groups = {}, {}
    for p in ok:
        if p["title"]:
            title_groups.setdefault(p["title"], []).append(p["url"])
        if p["meta_desc"]:
            meta_groups.setdefault(p["meta_desc"], []).append(p["url"])
    title_dupes = {t: urls for t, urls in title_groups.items() if len(urls) > 1}
    meta_dupes = {m: urls for m, urls in meta_groups.items() if len(urls) > 1}

    entries = []
    for t, urls in title_dupes.items():
        entries.append(("Critical", "Duplicate title", t, "\n".join(urls),
                        "Give each URL a unique, page-specific title. If pages are near-identical, canonical to the best version."))
    for m, urls in meta_dupes.items():
        entries.append(("High", "Duplicate meta description", m[:120] + "…", "\n".join(urls),
                        "Write unique meta description per URL."))

    start = 3
    for i, (pri, itype, shared, urls, fix) in enumerate(entries):
        r = start + i
        vals = [i + 1, "To Do", pri, itype, shared, urls, fix]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
        pc = ws.cell(row=r, column=3); pc.fill = PRI_FILL[pri]; pc.font = pri_font(pri)
        ws.row_dimensions[r].height = max(60, len(urls.split("\n")) * 15)
    end = start + len(entries) - 1
    if entries:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 14, 10, 22, 55, 65, 50])
    ws.freeze_panes = "C3"


def build_image_issues(wb, ok: list[dict]) -> None:
    ws = wb.create_sheet("Image Issues")
    headers = ["#", "Status", "URL", "Images missing alt", "Fix"]
    issues = [(p["url"], p["images_missing_alt"]) for p in ok if p["images_missing_alt"] > 0]
    add_title_row(ws, f"Image Issues — {len(issues)} pages with missing alt text", len(headers))
    add_header_row(ws, 2, headers)
    start = 3
    for i, (u, n) in enumerate(issues):
        r = start + i
        vals = [i + 1, "To Do", u, n, "Add descriptive alt text to every <img>. Use alt='' for purely decorative images."]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
    end = start + len(issues) - 1
    if issues:
        apply_list_validation(ws, "B", start, end, STATUS_VALUES)
        apply_status_cf(ws, "B", start, end, "A", get_column_letter(len(headers)))
    set_col_widths(ws, [5, 14, 75, 18, 65])


def build_gsc_top_pages(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("GSC - Top Pages")
    headers = ["Rank", "URL", "Clicks", "Impressions", "CTR (%)", "Avg Position"]
    add_title_row(ws, "Google Search Console — Top Pages", len(headers))
    add_header_row(ws, 2, headers)
    start = 3
    for i, rec in enumerate(rows):
        r = start + i
        vals = [i + 1, rec.get("url", rec.get("page")), int(rec.get("clicks", 0)),
                int(rec.get("impressions", 0)), float(rec.get("ctr", 0)),
                float(rec.get("position", 0))]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY
        if vals[4] < 1.0 and vals[3] > 20000:
            ws.cell(row=r, column=5).fill = FILL_HIGH; ws.cell(row=r, column=5).font = pri_font("High")
        if vals[5] > 20:
            ws.cell(row=r, column=6).fill = FILL_MED
    set_col_widths(ws, [6, 85, 10, 13, 10, 13])
    ws.freeze_panes = "C3"


def build_gsc_queries(wb, rows: list[dict], brand_regex: str | None) -> None:
    ws = wb.create_sheet("GSC - Top Queries")
    headers = ["Rank", "Query", "Clicks", "Impressions", "CTR (%)", "Avg Position", "Brand?"]
    add_title_row(ws, "Google Search Console — Top Queries", len(headers))
    add_header_row(ws, 2, headers)
    br = re.compile(brand_regex, re.I) if brand_regex else None
    start = 3
    for i, rec in enumerate(rows):
        r = start + i
        q = rec.get("query", "")
        is_brand = bool(br.search(q)) if br else False
        vals = [i + 1, q, int(rec.get("clicks", 0)), int(rec.get("impressions", 0)),
                float(rec.get("ctr", 0)), float(rec.get("position", 0)),
                "Yes" if is_brand else "No"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY
        if not is_brand and vals[4] < 1.0 and vals[3] > 5000:
            ws.cell(row=r, column=5).fill = FILL_HIGH; ws.cell(row=r, column=5).font = pri_font("High")
        if not is_brand and vals[5] > 10 and vals[3] > 5000:
            ws.cell(row=r, column=6).fill = FILL_MED
    set_col_widths(ws, [6, 45, 10, 14, 10, 13, 9])
    ws.freeze_panes = "C3"


def build_ahrefs_top_pages(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Ahrefs - Top Pages")
    headers = ["Rank", "URL", "Monthly traffic", "Traffic value (USD)", "Keywords", "Top keyword", "Ref domains"]
    add_title_row(ws, "Ahrefs — Top Pages by Organic Traffic", len(headers))
    add_header_row(ws, 2, headers)
    start = 3
    for i, rec in enumerate(rows):
        r = start + i
        val_cents = rec.get("value")
        val_usd = round((val_cents / 100), 2) if val_cents else 0
        vals = [i + 1, rec.get("url", ""), rec.get("sum_traffic", 0), val_usd,
                rec.get("keywords", 0), rec.get("top_keyword", ""),
                rec.get("referring_domains", 0)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
    set_col_widths(ws, [6, 85, 16, 18, 20, 38, 14])
    ws.freeze_panes = "C3"


def build_ahrefs_keywords(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Ahrefs - Keywords")
    headers = ["Rank", "Keyword", "Position", "Volume", "Traffic", "KD", "Branded?", "Ranking URL"]
    add_title_row(ws, "Ahrefs — Organic Keywords", len(headers))
    add_header_row(ws, 2, headers)
    start = 3
    for i, k in enumerate(rows):
        r = start + i
        vals = [i + 1, k.get("keyword", ""), k.get("best_position", 0),
                k.get("volume", 0), k.get("sum_traffic", 0),
                k.get("keyword_difficulty", ""),
                "Yes" if k.get("is_branded") else "No",
                k.get("best_position_url", "")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = BODY; c.alignment = WRAP
    set_col_widths(ws, [6, 45, 11, 11, 11, 8, 10, 70])
    ws.freeze_panes = "C3"


def build_references(wb) -> None:
    ws = wb.create_sheet("Google References")
    add_title_row(ws, "Google Official SEO References", 2)
    add_header_row(ws, 2, ["Topic", "URL"])
    for i, (t, u) in enumerate(REFS_TABLE):
        r = 3 + i
        a = ws.cell(row=r, column=1, value=t); a.border = BORDER; a.font = BODY; a.alignment = WRAP
        b = ws.cell(row=r, column=2, value=u); b.border = BORDER; b.font = BODY; b.alignment = WRAP
        b.hyperlink = u; b.style = "Hyperlink"
    set_col_widths(ws, [50, 90])


# ---------------------------- driver ----------------------------------------

def load_json(path: str | None) -> list | dict | None:
    if not path:
        return None
    with open(path) as f:
        d = json.load(f)
    if isinstance(d, dict):
        for key in ("rows", "pages", "keywords", "data", "results"):
            if key in d and isinstance(d[key], list):
                return d[key]
    return d


def main():
    ap = argparse.ArgumentParser(description="Build interactive SEO audit XLSX from crawl data.")
    ap.add_argument("--crawl", required=True, help="Path to crawl_results.json from crawl_site.py")
    ap.add_argument("--out", required=True, help="Output .xlsx path")
    ap.add_argument("--site-name", required=True, help="Human-readable site label (e.g. 'example.com/blog')")
    ap.add_argument("--gsc-top-pages", help="Optional JSON: list of {url/page, clicks, impressions, ctr, position}")
    ap.add_argument("--gsc-queries", help="Optional JSON: list of {query, clicks, impressions, ctr, position}")
    ap.add_argument("--gsc-devices", help="Optional JSON: list of {device, clicks, impressions, ctr, position}")
    ap.add_argument("--ahrefs-metrics", help="Optional JSON: dict with org_keywords / org_traffic")
    ap.add_argument("--ahrefs-top-pages", help="Optional JSON: list of top-pages records")
    ap.add_argument("--ahrefs-keywords", help="Optional JSON: list of keyword records")
    ap.add_argument("--dr", type=int, help="Domain Rating to show in Executive Summary")
    ap.add_argument("--brand-regex", help="Regex for marking queries as branded in GSC tab")
    args = ap.parse_args()

    crawl = json.load(open(args.crawl))
    ok = [p for p in crawl if p.get("status") == 200]
    print(f"Loaded {len(crawl)} pages ({len(ok)} status 200)", file=sys.stderr)

    gsc_top = load_json(args.gsc_top_pages)
    gsc_q = load_json(args.gsc_queries)
    gsc_dev = load_json(args.gsc_devices)
    ahrefs_metrics = load_json(args.ahrefs_metrics) if args.ahrefs_metrics else None
    ahrefs_top = load_json(args.ahrefs_top_pages)
    ahrefs_kws = load_json(args.ahrefs_keywords)

    gsc_totals = None
    if gsc_top:
        total_clk = sum(int(r.get("clicks", 0)) for r in gsc_top)
        total_imp = sum(int(r.get("impressions", 0)) for r in gsc_top)
        avg_ctr = (total_clk / total_imp * 100) if total_imp else 0
        avg_pos = sum(float(r.get("position", 0)) for r in gsc_top) / len(gsc_top)
        gsc_totals = {"clicks": total_clk, "impressions": total_imp, "ctr": avg_ctr, "position": avg_pos}

    scores = compute_scores(ok)

    wb = Workbook()
    wb.remove(wb.active)

    build_read_me(wb, args.site_name, len(ok))
    build_executive_summary(wb, args.site_name, ok, scores, ahrefs_metrics, gsc_totals, gsc_dev)
    build_action_plan(wb, ok, {"dr": args.dr})
    build_on_page_issues(wb, ok)
    build_schema_missing(wb, ok)
    build_hreflang_missing(wb, ok)
    build_thin_content(wb, ok)
    build_duplicate_titles_meta(wb, ok)
    build_image_issues(wb, ok)
    if gsc_top:
        build_gsc_top_pages(wb, gsc_top)
    if gsc_q:
        build_gsc_queries(wb, gsc_q, args.brand_regex)
    if ahrefs_top:
        build_ahrefs_top_pages(wb, ahrefs_top)
    if ahrefs_kws:
        build_ahrefs_keywords(wb, ahrefs_kws)
    build_page_inventory(wb, ok)
    build_references(wb)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb.save(args.out)
    print(f"Saved: {args.out}", file=sys.stderr)
    print(f"Sheets: {[s.title for s in wb.worksheets]}", file=sys.stderr)


if __name__ == "__main__":
    main()
